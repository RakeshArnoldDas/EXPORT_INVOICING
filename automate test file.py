from tkinter import *
import customtkinter
from CTkListbox import *
import pandas as pd
import openpyxl


customtkinter.set_appearance_mode("dark")
customtkinter.set_default_color_theme("blue")

root = customtkinter.CTk()
root.title('Tkinter.com - CustomTkinter ComboBox')
root.geometry('1920x1080')

import pandas as pd
import openpyxl


consignee_data_checkbox = None

# Open the Excel file for writing
invoice_file_path = 'I:/invoice/Invoice.xlsx'
invoice_wb = openpyxl.load_workbook(invoice_file_path)
invoice_sheet = invoice_wb.active




#NEW DATA ADDITION FUNCTIONS____________________________________________________________________________________________________________________________________________________________________________________________________________
def load_seller_details():

        import pandas as pd
        

        # Load SELLER names from Excel file
        excel_file_path = 'I:/databases/Seller details.xlsx'
        df = pd.read_excel(excel_file_path)
        company_names = df['Company name'].tolist()
        
        print(company_names)

        return df, company_names

def update_seller_names():
    global company_names
    _, company_names = load_seller_details()
    my_combo['values'] = company_names
    my_combo.configure(values=company_names)
    print("variable updated")

def load_buyer_details():

        # Load BUYER names from Excel file
        excel_file_path_buyer = 'I:/databases/Buyer details.xlsx'

        try:
            df_buyer = pd.read_excel(excel_file_path_buyer)
            buyer_names = df_buyer['Company name'].tolist()
        except FileNotFoundError:
            print(f"Error: File not found at path: {excel_file_path_buyer}")
            buyer_names = []
        except Exception as e:
            print(f"Error loading data from '{excel_file_path_buyer}': {e}")
            buyer_names = []

        # Check if 'Company name' column exists
        if 'Company name' not in df_buyer.columns:
            print("Error: 'Company name' column not found in the 'Buyer details.xlsx' file.")
            buyer_names = []

        return df_buyer, buyer_names
        

def update_buyer_names():
    global buyer_names
    _, buyer_names = load_buyer_details()
    my_comboseller['values'] = buyer_names
    my_comboseller.configure(values=buyer_names)
    print("variable updated")



def load_Notify_details():
       # Load NOTIFY names from the Excel file
        excel_file_path_notify = 'I:/databases/Notify details.xlsx'
        df_notify = pd.read_excel(excel_file_path_notify)

        if df_notify.empty:
            print("Error: DataFrame 'df_notify' is empty.")
        else:
            # Get all company names, including duplicates
            notify_company_names = df_notify['Company name'].tolist()

        return df_notify, notify_company_names

def update_notify_names():
    global notify_company_names
    _, notify_company_names = load_Notify_details()
    my_combo_notify['values'] = notify_company_names
    my_combo_notify.configure(values=notify_company_names)
    print("variable updated")







#_____________________________________________________________________________________________________________________________________________________________________________________________________________________________________

# Seller details
def choosecompany():

    df, company_names = load_seller_details()
    selected_company = my_combo.get()

    if not df.empty:
        selected_company_details = df[df['Company name'] == selected_company].iloc[0]
        # Write data to the specified cells in the Invoice.xlsx sheet        
        invoice_sheet['A16'] = selected_company_details['Company name']
        invoice_sheet['A17'] = f"Attn: {selected_company_details['Contact person']}"
        invoice_sheet['A18'] = f"{selected_company_details['Address line 1']} {selected_company_details['Address line 2']}"
        invoice_sheet['A19'] = f"{selected_company_details['City']}, {selected_company_details['State']} {selected_company_details['Pin code']}, {selected_company_details['Country']}"
        invoice_sheet['A20'] = f"Phn No: {selected_company_details['Phn No']}"

        # Additional data
        invoice_sheet['B38'] = selected_company_details.get('GST No', '')
        invoice_sheet['B39'] = selected_company_details.get('State code', '')
        iec_no_text = f"IEC No: {selected_company_details.get('IEC No', '')}"
        ad_code_text = f"AD Code: {selected_company_details.get('AD code', '')}"
        br_code_text = f"BR Code: {selected_company_details.get('Br code', '')}"
        invoice_sheet['A37'] = f"{iec_no_text}, {ad_code_text}, {br_code_text}"
    else:
         print("Error: DataFrame 'df' is empty.")

    # Save the changes to the Excel file
    invoice_wb.save(invoice_file_path)

    # Update the output label with the concatenated address
    concatenated_address = ' '.join(
        f"{('Attn:' if col == 'Contact person' else 'Phn No:' if col == 'Phn No' else '')} {selected_company_details[col]}"
        for col in ['Company name', 'Contact person', 'Address line 1', 'Address line 2', 'City', 'State', 'Pin code', 'Country', 'Phn No']
        if not pd.isna(selected_company_details[col])
    )
    output_label.configure(text=concatenated_address, text_color="white")





# Buyer details
def choosebuyer():

    df_buyer, buyer_names = load_buyer_details()
    selected_buyer = my_comboseller.get()
    print(selected_buyer)
    selected_buyer_details = df_buyer[df_buyer['Company name'] == selected_buyer].iloc[0]

    # Open the Excel file for writing
    invoice_wb = openpyxl.load_workbook(invoice_file_path)
    invoice_sheet = invoice_wb.active

    # Concatenate the buyer details
    concatenated_buyer_details = ' '.join(
        f"{('Attn:' if col == 'Contact person' else 'Phn No:' if col == 'Phn No' else '')} {selected_buyer_details[col]}"
        for col in ['Company name', 'Contact person', 'Address line 1', 'Address line 2', 'City', 'State', 'Pin code', 'Country', 'Phn No']
        if not pd.isna(selected_buyer_details[col])
    )

    # Display the concatenated buyer details in the output_label_buyer
    output_label_buyer.configure(text=concatenated_buyer_details, text_color="white")

    # Write data to the specified cells in the Invoice.xlsx sheet
    invoice_sheet['E20'] = concatenated_buyer_details

    # Save the changes to the Excel file
    invoice_wb.save(invoice_file_path)


# Consignee details
def getdatatest1():
    if consignee_data_checkbox is not None:
        print("Data retrieved:", consignee_data_checkbox)
    else:
        print("No data retrieved from popup.")







def consignee_details():
    

    df_buyer, _ = load_buyer_details()
    # Get the text from output_label_buyer
    buyer_details_text = output_label_buyer.cget("text")

    # Set the consignee_label text to the buyer details text
    consignee_label.configure(text="Buyer is same as Consignee: " + buyer_details_text)

    # Disable or enable the consignee_details button based on consignee_check state
    if check_var.get() == "on":
        my_button_consignee_details.configure(state="disabled")
    else:
        my_button_consignee_details.configure(state="normal")

    selected_buyer = my_comboseller.get()
    selected_buyer_details = df_buyer[df_buyer['Company name'] == selected_buyer].iloc[0]

    # Write data to the specified cells in the Invoice.xlsx sheet
    invoice_sheet['A22'] = selected_buyer_details['Company name']
    invoice_sheet['A23'] = f"Attn: {selected_buyer_details['Contact person']}"
    invoice_sheet['A24'] = f"{selected_buyer_details['Address line 1']} "
    invoice_sheet['A25'] = f"{selected_buyer_details['Address line 2']}"
    invoice_sheet['A26'] = f"{selected_buyer_details['City']}, {selected_buyer_details['State']} {selected_buyer_details['Pin code']}"
    invoice_sheet['A27'] = f"{selected_buyer_details['Country']}"
    invoice_sheet['A28'] = f"Phn No: {selected_buyer_details['Phn No']}"

    # Save the changes to the Excel file
    invoice_wb.save(invoice_file_path)


def list_df_consignee():
    # Load CONSIGNEE names from Excel file 
    excel_file_path1 = 'I:/databases/Consignee details.xlsx'
    df = pd.read_excel(excel_file_path1)
    df.fillna('NA', inplace=True)
    company_names = df['Company name'].tolist()
    print(company_names)
    return df

#Consignee as second notify party
def secondconsignee_details():

    invoice_file_path = 'I:/invoice/Invoice.xlsx'
    invoice_wb = openpyxl.load_workbook(invoice_file_path)
    invoice_sheet = invoice_wb.active


    if check_var2.get() == "on":
        print("Value of check_var2:", check_var2.get())


        df_buyer, _ = load_buyer_details()

        # Get the text from output_label_buyer
        buyer_details_text = output_label_buyer.cget("text")

        # Set the consignee_label text to the buyer details text
        secondconsignee_label.configure(text="Second Notify Party: " + buyer_details_text)


        selected_buyer = my_comboseller.get()
        selected_buyer_details = df_buyer[df_buyer['Company name'] == selected_buyer].iloc[0]

        # Write data to the specified cells in the Invoice.xlsx sheet
        invoice_sheet['H24'] = selected_buyer_details['Company name']
        invoice_sheet['H25'] = f"Attn: {selected_buyer_details['Contact person']}"
        invoice_sheet['H26'] = f"{selected_buyer_details['Address line 1']} "
        invoice_sheet['H27'] = f"{selected_buyer_details['Address line 2']}"
        invoice_sheet['H28'] = f"{selected_buyer_details['City']}, {selected_buyer_details['State']} {selected_buyer_details['Pin code']}"
        invoice_sheet['H29'] = f"{selected_buyer_details['Country']}"
        invoice_sheet['H30'] = f"Phn No: {selected_buyer_details['Phn No']}"

        # Save the changes to the Excel file
        invoice_wb.save(invoice_file_path)
        
    else:
        df = list_df_consignee()
        print("Value of check_var2:", check_var2.get())


        # Find corresponding entry in the DataFrame
        selected_row = df[df['Company name'] == consignee_data_checkbox].iloc[0]

        # Concatenate all rows, ignoring blank cells
        concatenated_details = ""
        for column, value in selected_row.items():
            if value != 'NA':
                concatenated_details = f"{selected_row['City']}, {selected_row['State']} {selected_row['Pin code']},{selected_row['Country']}"
        

        company_name = selected_row['Company name']
        contact_person = selected_row['Contact person']
        address1 = selected_row['Address line 1']
        address2 = selected_row['Address line 2']
        address3 = selected_row['Address line 3']
        citystatepincountry = concatenated_details
        country = selected_row['Country']
        phone = selected_row['Phn No']        

        print(company_name)
        print(contact_person )
        print(  address1 )
        print(address2 )
        print( address3 )
        print(  citystatepincountry )
        print(  country )
        print(        phone)


        # Write data to the specified cells in the Invoice.xlsx sheet
        invoice_sheet['H24'] = (company_name )
        invoice_sheet['H25'] = (contact_person)
        invoice_sheet['H26'] = (address1)
        invoice_sheet['H27'] = ((address2 + address3))
        invoice_sheet['H28'] = (citystatepincountry)
        invoice_sheet['H29'] = (country)
        invoice_sheet['H30'] = (f"Phn No: {phone}")

        # Save the changes to the Excel file
        invoice_wb.save(invoice_file_path)

        # Convert all variables to strings
        company_name_str = str(company_name)
        contact_person_str = str(contact_person)
        address1_str = str(address1)
        address2_str = str(address2)
        address3_str = str(address3)
        citystatepincountry_str = str(citystatepincountry)
        phone_str = str(phone)

        # Concatenate all strings into one
        message = company_name_str + " Attn: " + contact_person_str + " ," + address1_str + " ," + address2_str + " ," + address3_str + " ," + citystatepincountry_str + " Phn No: " + phone_str

        # Configure consignee_label with the concatenated string
        secondconsignee_label.configure(text="Second Notify Party: " + message)


# Notify party details
def choosenotify():

    df_notify, notify_company_names = load_Notify_details()

    selected_notify = my_combo_notify.get()
    selected_notify_details = df_notify[df_notify['Company name'] == selected_notify].iloc[0]

    # Write data to the specified cells in the Invoice.xlsx sheet
    # Write data to the specified cells in the Invoice.xlsx sheet
    invoice_sheet['E24'] = selected_notify_details['Company name']
    invoice_sheet['E25'] = f"Attn: {selected_notify_details['Contact person']}"
    invoice_sheet['E26'] = f"{selected_notify_details['Address line 1']} "
    invoice_sheet['E27'] = f"{selected_notify_details['Address line 2']}{selected_notify_details['Address line 3']}"
    invoice_sheet['E28'] = f"{selected_notify_details['City']}, {selected_notify_details['State']} {selected_notify_details['Pin code']}"
    invoice_sheet['E29'] = f"{selected_notify_details['Country']}"
    invoice_sheet['E30'] = f"Phn No: {selected_notify_details['Phn No']}"
    # Save the changes to the Excel file
    invoice_wb.save(invoice_file_path)
    print("Data saved successfully.")

    # Update the output label with the concatenated address
    concatenated_address = ' '.join(
        f"{('Attn:' if col == 'Contact person' else 'Phn No:' if col == 'Phn No' else '')} {selected_notify_details[col]}"
        for col in ['Company name', 'Contact person', 'Address line 1', 'Address line 2', 'City', 'State', 'Pin code', 'Country', 'Phn No']
        if not pd.isna(selected_notify_details[col])
    )
    output_label_notify.configure(text=concatenated_address, text_color="white")





# Port of Discharge
def submitport():
    port_value = portofdelivery.get()  # Retrieve the text value from the entry widget
    portofdeliverylabel.configure(text=f'Port of delivery: ' + port_value)

    # Specify the path to your Excel file
    excel_file_path = 'I:/invoice/Invoice.xlsx'

    # Load the workbook
    workbook = openpyxl.load_workbook(excel_file_path)

    # Select the default active sheet (assuming it's the first sheet)
    sheet = workbook.active

    # Write the value to cell C30
    sheet['C30'] = port_value  # Use port_value instead of portofdelivery

    # Save the changes
    workbook.save(excel_file_path)

       

def get_rad():

    if incoterms_var.get() == "other":
        incoterms_label.configure(text="Please Make A Selection")
    elif incoterms_var.get() == "Yes":
        incoterms_label.configure(text=f'INCOTERMS : '+"CIF")
    else:
        incoterms_label.configure(text=f'INCOTERMS : '+"FOB")


    # Convert the value to a string before writing to cell C31
    value_to_write = str(incoterms_label.cget("text"))

    # Specify the path to your Excel file
    excel_file_path = 'I:/invoice/Invoice.xlsx'

    # Load the workbook
    workbook = openpyxl.load_workbook(excel_file_path)

    # Select the default active sheet (assuming it's the first sheet)
    sheet = workbook.active

    # Write the value to cell C30
    sheet['C31'] = value_to_write  

    # Save the changes
    workbook.save(excel_file_path)

def shippingterms_get():

    if shippingmode_var.get() == 1:
        shippingmode_label.configure(text="Please Make A Selection")
    elif shippingmode_var.get() == 2:
        shippingmode_label.configure(text= "SEA FREIGHT")
    elif shippingmode_var.get() == 3:
        shippingmode_label.configure(text="AIR FREIGHT")
    else:
        shippingmode_label.configure(text="COURIER")


    # Convert the value to a string before writing to cell C31
    shippingvalue_to_write = str(shippingmode_label.cget("text"))

    # Specify the path to your Excel file
    excel_file_path = 'I:/invoice/Invoice.xlsx'

    # Load the workbook
    workbook = openpyxl.load_workbook(excel_file_path)

    # Select the default active sheet (assuming it's the first sheet)
    sheet = workbook.active

    # Write the value to cell C30
    sheet['C34'] = shippingvalue_to_write  

    # Save the changes
    workbook.save(excel_file_path)

def payment_terms():
    if shippingmode_var.get() == 1:
        payment=str("ERROR")
    elif shippingmode_var.get() ==2:
        payment=str("90 days from delivery")
    elif shippingmode_var == 3:
        payment=str("90 days from delivery")
    else:
        payment=str("30 days from delivery")

    # Specify the path to your Excel file
    excel_file_path = 'I:/invoice/Invoice.xlsx'

    # Load the workbook
    workbook = openpyxl.load_workbook(excel_file_path)

    # Select the default active sheet (assuming it's the first sheet)
    sheet = workbook.active

    # Write the value to cell C30
    sheet['C32'] = payment

    # Save the changes
    workbook.save(excel_file_path)

def shippingtermpayment():
        shippingterms_get()
        payment_terms()



def country(port_value):
    # Sample list of countries (you can replace it with your own list)
    countries = ['AFGHANISTAN', 'ALBANIA', 'ALGERIA', 'ANDORRA', 'ANGOLA', 'ANTIGUA AND BARBUDA', 'ARGENTINA', 'ARMENIA', 'AUSTRALIA', 'AUSTRIA', 'AZERBAIJAN', 'BAHAMAS', 'BAHRAIN', 'BANGLADESH', 'BARBADOS', 'BELARUS', 'BELGIUM', 'BELIZE', 'BENIN', 'BHUTAN', 'BOLIVIA', 'BOSNIA AND HERZEGOVINA', 'BOTSWANA', 'BRAZIL', 'BRUNEI', 'BULGARIA', 'BURKINA FASO', 'BURUNDI', 'CABO VERDE', 'CAMBODIA', 'CAMEROON', 'CANADA', 'CENTRAL AFRICAN REPUBLIC', 'CHAD', 'CHILE', 'CHINA', 'COLOMBIA', 'COMOROS', 'CONGO (CONGO-BRAZZAVILLE)', 'COSTA RICA', 'CROATIA', 'CUBA', 'CYPRUS', 'CZECHIA (CZECH REPUBLIC)', 'DEMOCRATIC REPUBLIC OF THE CONGO', 'DENMARK', 'DJIBOUTI', 'DOMINICA', 'DOMINICAN REPUBLIC', 'ECUADOR', 'EGYPT', 'EL SALVADOR', 'EQUATORIAL GUINEA', 'ERITREA', 'ESTONIA', 'ESWATINI (FMR. "SWAZILAND")', 'ETHIOPIA', 'FIJI', 'FINLAND', 'FRANCE', 'GABON', 'GAMBIA', 'GEORGIA', 'GERMANY', 'GHANA', 'GREECE', 'GRENADA', 'GUATEMALA', 'GUINEA', 'GUINEA-BISSAU', 'GUYANA', 'HAITI', 'HOLY SEE', 'HONDURAS', 'HUNGARY', 'ICELAND', 'INDIA', 'INDONESIA', 'IRAN', 'IRAQ', 'IRELAND', 'ISRAEL', 'ITALY', 'JAMAICA', 'JAPAN', 'JORDAN', 'KAZAKHSTAN', 'KENYA', 'KIRIBATI', 'KUWAIT', 'KYRGYZSTAN', 'LAOS', 'LATVIA', 'LEBANON', 'LESOTHO', 'LIBERIA', 'LIBYA', 'LIECHTENSTEIN', 'LITHUANIA', 'LUXEMBOURG', 'MADAGASCAR', 'MALAWI', 'MALAYSIA', 'MALDIVES', 'MALI', 'MALTA', 'MARSHALL ISLANDS', 'MAURITANIA', 'MAURITIUS', 'MEXICO', 'MICRONESIA', 'MOLDOVA', 'MONACO', 'MONGOLIA', 'MONTENEGRO', 'MOROCCO', 'MOZAMBIQUE', 'MYANMAR (FORMERLY BURMA)', 'NAMIBIA', 'NAURU', 'NEPAL', 'NETHERLANDS', 'NEW ZEALAND', 'NICARAGUA', 'NIGER', 'NIGERIA', 'NORTH KOREA', 'NORTH MACEDONIA (FORMERLY MACEDONIA)', 'NORWAY', 'OMAN', 'PAKISTAN', 'PALAU', 'PALESTINE STATE', 'PANAMA', 'PAPUA NEW GUINEA', 'PARAGUAY', 'PERU', 'PHILIPPINES', 'POLAND', 'PORTUGAL', 'QATAR', 'ROMANIA', 'RUSSIA', 'RWANDA', 'SAINT KITTS AND NEVIS', 'SAINT LUCIA', 'SAINT VINCENT AND THE GRENADINES', 'SAMOA', 'SAN MARINO', 'SAO TOME AND PRINCIPE', 'SAUDI ARABIA', 'SENEGAL', 'SERBIA', 'SEYCHELLES', 'SIERRA LEONE', 'SINGAPORE', 'SLOVAKIA', 'SLOVENIA', 'SOLOMON ISLANDS', 'SOMALIA', 'SOUTH AFRICA', 'SOUTH KOREA', 'SOUTH SUDAN', 'SPAIN', 'SRI LANKA', 'SUDAN', 'SURINAME', 'SWEDEN', 'SWITZERLAND', 'SYRIA', 'TAJIKISTAN', 'TANZANIA', 'THAILAND', 'TIMOR-LESTE', 'TOGO', 'TONGA', 'TRINIDAD AND TOBAGO', 'TUNISIA', 'TURKEY', 'TURKMENISTAN', 'TUVALU', 'UGANDA', 'UKRAINE', 'UNITED ARAB EMIRATES', 'UNITED KINGDOM', 'UNITED STATES OF AMERICA', 'URUGUAY', 'UZBEKISTAN', 'VANUATU', 'VENEZUELA', 'VIETNAM', 'YEMEN', 'ZAMBIA', 'ZIMBABWE', 'USA']

    # Input string provided by the user
    # Sample input string
    port_value = port_value.strip()

    print(f"The port value entered by user is: '{port_value}'")

    # Extracting the last word
    last_word = port_value.split(",")[-1].strip()

    print(f"The extracted last word is: '{last_word}'")  # Debugging line

    # Checking if the last word is a country
    if last_word in countries:
        # Convert the value to a string before writing to cell C31
        country_to_write = str(last_word)

        # Specify the path to your Excel file
        excel_file_path = 'I:/invoice/Invoice.xlsx'

        # Load the workbook
        workbook = openpyxl.load_workbook(excel_file_path)

        # Select the default active sheet (assuming it's the first sheet)
        sheet = workbook.active

        # Write the value to cell I32
        sheet['I32'] = country_to_write

        # Save the changes
        workbook.save(excel_file_path)
    else:
        print(f"The last word '{last_word}' is not a valid country.")

            
def submit_and_country():
    port_value = portofdelivery.get()
    submitport()
    country(port_value)

def shippingmarksinput():
        marks_value=shippingmarks.get()
        marks_label.configure(text=f'SHIPPING MARKS: ' + marks_value)

        # Convert the value to a string before writing to cell C31
        marks_to_write = str(marks_value)

        # Specify the path to your Excel file
        excel_file_path = 'I:/invoice/Invoice.xlsx'

        # Load the workbook
        workbook = openpyxl.load_workbook(excel_file_path)

        # Select the default active sheet (assuming it's the first sheet)
        sheet = workbook.active

        # Write the value to cell I32
        sheet['C35'] = marks_to_write

        # Save the changes
        workbook.save(excel_file_path)


#NEW SELLER DATA ADDITION FUNCTION_____________________________________________________________________________________________________________________________________________________________________________________________________
def new_seller():
    newseller_window  = customtkinter.CTkToplevel(root)
    newseller_window.title("New Seller")
    newseller_window.geometry("400x800")
    newseller_window.attributes(('-topmost'),True)
    newseller_window.grab_set()

    

    import pandas as pd

        # Destroy the window after a delay
    def newseller_window_destroy():    
        newseller_window.after(1500, newseller_window.destroy)

    def sellerdatarow():
        
        companyname = companyname_seller.get()
        contactperson = contactperson_seller.get()
        address1 = addressline1_seller.get()
        address2 = Addressline2_seller.get()
        City = City_seller.get()
        State = State_seller.get()
        pincode = pincode_seller.get()
        Country = Country_seller.get()
        Phn_No = Phn_No_seller.get()
        gst_no = gst_no_seller.get()
        iec_no = iec_no_seller.get()
        adcode = adcode_no_seller.get()
        Statecode = Statecode_no_seller.get()
        brcode = brcode_no_seller.get()

        # Return a dictionary with the values
        return {

            "Company name": companyname,
            "Contact person": contactperson,
            "Address line 1": address1,
            "Address line 2": address2,
            "City": City,
            "State": State,
            "Pin code": pincode,
            "Country": Country,
            "Phn No": Phn_No,
            "GST No": gst_no,
            "IEC No": iec_no,
            "AD code": adcode,
            "State code": Statecode,
            "Br code": brcode
        }

    # Function to run the script and save data
    def run_script_and_save():
        try:
            import openpyxl
        except ImportError:
            print("Error", "openpyxl module not found. Make sure it's installed.")
            return

        def find_first_row_with_data(sheet, start_row):
            current_row = start_row
            while sheet.cell(row=current_row, column=1).value is None:
                current_row -= 1
            return current_row

        # Main function to handle data mapping and writing to Excel using Pandas
        def main():
            data_dict = sellerdatarow()
            print(data_dict)

            # Define column mapping with correct capitalization
            column_mapping = {
                "Company name": "Company name",
                "Contact person": "Contact person",
                "Address line 1": "Address line 1",
                "Address line 2": "Address line 2",
                "City": "City",
                "State": "State",
                "Pin code": "Pin code",
                "Country": "Country",
                "Phn No": "Phn No",
                "GST No": "GST No",
                "IEC No": "IEC No",
                "AD code": "AD code",
                "State code": "State code",
                "Br code": "Br code"
            }

            print("Column Mapping Keys:", column_mapping.keys())

            # Map data inputs to column names
            data_mapped = {column_mapping[key]: value for key, value in data_dict.items()if key in column_mapping}
            print("Mapped data:", data_mapped)

            # Create a DataFrame with the mapped data
            df = pd.DataFrame([data_mapped])

            # Load existing data from Excel
            try:
                existing_data = pd.read_excel('I:/databases/Seller details.xlsx')
            except FileNotFoundError:
                print("ERROR FILE PATH IS INCORRECT")
                existing_data = pd.DataFrame()

            # Concatenate existing data with new data
            updated_data = pd.concat([existing_data, df], ignore_index=True)

            # Write back to Excel
            updated_data.to_excel('I:/databases/Seller details.xlsx', index=False,)
            print("Success", "Data entered successfully!")

            # Update label for user:
            newsellerstatus_label.configure(text=f'SELLER DETAILS ENTERED SUCCESSFULLY')

            # Reload the data for the rest of the code
            load_seller_details()

            #update the combobox:
            update_seller_names()            

            # Destroy the window after a delay
            newseller_window_destroy()

        main()


    companyname_seller = customtkinter.CTkEntry(newseller_window, placeholder_text="Company name")
    companyname_seller.grid(row=0,column=1,padx=10,pady=5,sticky=W,columnspan=2)
    
    contactperson_seller = customtkinter.CTkEntry(newseller_window, placeholder_text="Contact Person details")
    contactperson_seller.grid(row=1,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    addressline1_seller = customtkinter.CTkEntry(newseller_window, placeholder_text="Address Line 1")
    addressline1_seller.grid(row=2,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    Addressline2_seller = customtkinter.CTkEntry(newseller_window, placeholder_text="Address Line 2")
    Addressline2_seller.grid(row=3,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    City_seller = customtkinter.CTkEntry(newseller_window, placeholder_text="City")
    City_seller.grid(row=4,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    State_seller = customtkinter.CTkEntry(newseller_window, placeholder_text="State")
    State_seller.grid(row=5,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    pincode_seller = customtkinter.CTkEntry(newseller_window, placeholder_text="Pin code")
    pincode_seller.grid(row=6,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    Country_seller = customtkinter.CTkEntry(newseller_window, placeholder_text="Country")
    Country_seller.grid(row=7,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    Phn_No_seller = customtkinter.CTkEntry(newseller_window, placeholder_text="Phn No")
    Phn_No_seller.grid(row=8,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    gst_no_seller = customtkinter.CTkEntry(newseller_window, placeholder_text="GST No")
    gst_no_seller.grid(row=9,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    iec_no_seller = customtkinter.CTkEntry(newseller_window, placeholder_text="IEC No")
    iec_no_seller.grid(row=10,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    adcode_no_seller = customtkinter.CTkEntry(newseller_window, placeholder_text="AD code No")
    adcode_no_seller.grid(row=11,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    Statecode_no_seller = customtkinter.CTkEntry(newseller_window, placeholder_text="State code")
    Statecode_no_seller.grid(row=12,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    
    brcode_no_seller = customtkinter.CTkEntry(newseller_window, placeholder_text="BR code")
    brcode_no_seller.grid(row=13,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    save_button = customtkinter.CTkButton(newseller_window,text="Save",command=run_script_and_save)
    save_button.grid(row=16,column=1,padx=10,pady=5,sticky=W)

    discard_button = customtkinter.CTkButton(newseller_window,text="Discard",command=newseller_window_destroy)
    discard_button.grid(row=18,column=1,padx=10,pady=5,sticky=W)

    newsellerstatus_label = customtkinter.CTkLabel(newseller_window,text="",font=("Tahoma",12))
    newsellerstatus_label.grid(row=19,column=1,padx=10,pady=5,sticky=W)


#NEW BUYER DATA ADDITION FUNCTION______________________________________________________________________________________________________________________________________________________________________________________________________________________
def new_buyer():
    newbuyer_window  = customtkinter.CTkToplevel(root)
    newbuyer_window.title("New Buyer")
    newbuyer_window.geometry("400x800")
    newbuyer_window.attributes(('-topmost'),True)
    newbuyer_window.grab_set()

    

    import pandas as pd

        # Destroy the window after a delay
    def new_buyer_window_destroy():    
        newbuyer_window.after(1500, newbuyer_window.destroy)

    def buyerdatarow():
        
        companyname = companyname_buyer.get()
        contactperson = contactperson_buyer.get()
        address1 = addressline1_buyer.get()
        address2 = Addressline2_buyer.get()
        address3 = Addressline3_buyer.get()
        City = City_buyer.get()
        State = State_buyer.get()
        pincode = pincode_buyer.get()
        Country = Country_buyer.get()
        Phn_No = Phn_No_buyer.get()
        

        # Return a dictionary with the values
        return {

            "Company name": companyname,
            "Contact person": contactperson,
            "Address line 1": address1,
            "Address line 2": address2,
            "Address line 3": address3,
            "City": City,
            "State": State,
            "Pin code": pincode,
            "Country": Country,
            "Phn No": Phn_No,
            
        }

    # Function to run the script and save data
    def run_buyer_script_and_save():
        try:
            import openpyxl
        except ImportError:
            print("Error", "openpyxl module not found. Make sure it's installed.")
            return

        def find_first_row_with_data(sheet, start_row):
            current_row = start_row
            while sheet.cell(row=current_row, column=1).value is None:
                current_row -= 1
            return current_row

        # Main function to handle data mapping and writing to Excel using Pandas
        def main_buyer():
            data_dict = buyerdatarow()
            print(data_dict)

            # Define column mapping with correct capitalization
            column_mapping = {
                "Company name": "Company name",
                "Contact person": "Contact person",
                "Address line 1": "Address line 1",
                "Address line 2": "Address line 2",
                "Adderss line 3": "Address line 3",
                "City": "City",
                "State": "State",
                "Pin code": "Pin code",
                "Country": "Country",
                "Phn No": "Phn No",
 
            }

            print("Column Mapping Keys:", column_mapping.keys())

            # Map data inputs to column names
            data_mapped = {column_mapping[key]: value for key, value in data_dict.items()if key in column_mapping}
            print("Mapped data:", data_mapped)

            # Create a DataFrame with the mapped data
            df = pd.DataFrame([data_mapped])

            # Load existing data from Excel
            try:
                existing_data = pd.read_excel('I:/databases/Buyer details.xlsx')
            except FileNotFoundError:
                print("ERROR FILE PATH IS INCORRECT")
                existing_data = pd.DataFrame()

            # Concatenate existing data with new data
            updated_data = pd.concat([existing_data, df], ignore_index=True)

            # Write back to Excel
            updated_data.to_excel('I:/databases/Buyer details.xlsx', index=False,)
            print("Success", "Data entered successfully!")

            # Update label for user:
            newbuyerstatus_label.configure(text=f'BUYER DETAILS ENTERED SUCCESSFULLY')

            # Reload the data for the rest of the code
            load_buyer_details()

            #update the combobox:
            update_buyer_names()            

            # Destroy the window after a delay
            new_buyer_window_destroy()

        main_buyer()


    companyname_buyer = customtkinter.CTkEntry(newbuyer_window, placeholder_text="Company name")
    companyname_buyer.grid(row=0,column=1,padx=10,pady=5,sticky=W,columnspan=2)
    
    contactperson_buyer = customtkinter.CTkEntry(newbuyer_window, placeholder_text="Contact Person details")
    contactperson_buyer.grid(row=1,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    addressline1_buyer = customtkinter.CTkEntry(newbuyer_window, placeholder_text="Address Line 1")
    addressline1_buyer.grid(row=2,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    Addressline2_buyer = customtkinter.CTkEntry(newbuyer_window, placeholder_text="Address Line 2")
    Addressline2_buyer.grid(row=3,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    Addressline3_buyer = customtkinter.CTkEntry(newbuyer_window, placeholder_text="Address Line 3")
    Addressline3_buyer.grid(row=4,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    City_buyer = customtkinter.CTkEntry(newbuyer_window, placeholder_text="City")
    City_buyer.grid(row=5,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    State_buyer = customtkinter.CTkEntry(newbuyer_window, placeholder_text="State")
    State_buyer.grid(row=6,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    pincode_buyer = customtkinter.CTkEntry(newbuyer_window, placeholder_text="Pin code")
    pincode_buyer.grid(row=7,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    Country_buyer = customtkinter.CTkEntry(newbuyer_window, placeholder_text="Country")
    Country_buyer.grid(row=8,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    Phn_No_buyer = customtkinter.CTkEntry(newbuyer_window, placeholder_text="Phn No")
    Phn_No_buyer.grid(row=9,column=1,padx=10,pady=5,sticky=W,columnspan=2)






    save_button = customtkinter.CTkButton(newbuyer_window,text="Save",command=run_buyer_script_and_save)
    save_button.grid(row=17,column=1,padx=10,pady=5,sticky=W)

    discard_button = customtkinter.CTkButton(newbuyer_window,text="Discard",command=new_buyer_window_destroy)
    discard_button.grid(row=19,column=1,padx=10,pady=5,sticky=W)

    newbuyerstatus_label = customtkinter.CTkLabel(newbuyer_window,text="",font=("Tahoma",12))
    newbuyerstatus_label.grid(row=20,column=1,padx=10,pady=5,sticky=W)





#NEW NOTIFY PARTY DATA ADDITION FUNCTION__________________________________________________________________________________________________________________________________________________________________________________________________
def new_notify():
    newnotify_window  = customtkinter.CTkToplevel(root)
    newnotify_window.title("New Buyer")
    newnotify_window.geometry("400x800")
    newnotify_window.attributes(('-topmost'),True)
    newnotify_window.grab_set()

    

    import pandas as pd

        # Destroy the window after a delay
    def new_notify_window_destroy():    
        newnotify_window.after(1500, newnotify_window.destroy)

    def notifydatarow():
        
        companyname = companyname_notify.get()
        contactperson = contactperson_notify.get()
        address1 = addressline1_notify.get()
        address2 = Addressline2_notify.get()
        address3 = Addressline3_notify.get()
        City = City_notify.get()
        State = State_notify.get()
        pincode = pincode_notify.get()
        Country = Country_notify.get()
        Phn_No = Phn_No_notify.get()
        

        # Return a dictionary with the values
        return {

            "Company name": companyname,
            "Contact person": contactperson,
            "Address line 1": address1,
            "Address line 2": address2,
            "Address line 3": address3,
            "City": City,
            "State": State,
            "Pin code": pincode,
            "Country": Country,
            "Phn No": Phn_No,
            
        }

    # Function to run the script and save data
    def run_notify_script_and_save():
        try:
            import openpyxl
        except ImportError:
            print("Error", "openpyxl module not found. Make sure it's installed.")
            return

        def find_first_row_with_data(sheet, start_row):
            current_row = start_row
            while sheet.cell(row=current_row, column=1).value is None:
                current_row -= 1
            return current_row

        # Main function to handle data mapping and writing to Excel using Pandas
        def main_notify():
            data_dict = notifydatarow()
            print(data_dict)

            # Define column mapping with correct capitalization
            column_mapping = {
                "Company name": "Company name",
                "Contact person": "Contact person",
                "Address line 1": "Address line 1",
                "Address line 2": "Address line 2",
                "Adderss line 3": "Address line 3",
                "City": "City",
                "State": "State",
                "Pin code": "Pin code",
                "Country": "Country",
                "Phn No": "Phn No",
 
            }

            print("Column Mapping Keys:", column_mapping.keys())

            # Map data inputs to column names
            data_mapped = {column_mapping[key]: value for key, value in data_dict.items()if key in column_mapping}
            print("Mapped data:", data_mapped)

            # Create a DataFrame with the mapped data
            df = pd.DataFrame([data_mapped])

            # Load existing data from Excel
            try:
                existing_data = pd.read_excel('I:/databases/Notify details.xlsx')
            except FileNotFoundError:
                print("ERROR FILE PATH IS INCORRECT")
                existing_data = pd.DataFrame()

            # Concatenate existing data with new data
            updated_data = pd.concat([existing_data, df], ignore_index=True)

            # Write back to Excel
            updated_data.to_excel('I:/databases/Notify details.xlsx', index=False,)
            print("Success", "Data entered successfully!")

            # Update label for user:
            newnotifystatus_label.configure(text=f'NOTIFY DETAILS ENTERED SUCCESSFULLY')

            # Reload the data for the rest of the code
            load_Notify_details()

            #update the combobox:
            update_notify_names()            

            # Destroy the window after a delay
            new_notify_window_destroy()

        main_notify()


    companyname_notify = customtkinter.CTkEntry(newnotify_window, placeholder_text="Company name")
    companyname_notify.grid(row=0,column=1,padx=10,pady=5,sticky=W,columnspan=2)
    
    contactperson_notify = customtkinter.CTkEntry(newnotify_window, placeholder_text="Contact Person details")
    contactperson_notify.grid(row=1,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    addressline1_notify = customtkinter.CTkEntry(newnotify_window, placeholder_text="Address Line 1")
    addressline1_notify.grid(row=2,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    Addressline2_notify = customtkinter.CTkEntry(newnotify_window, placeholder_text="Address Line 2")
    Addressline2_notify.grid(row=3,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    Addressline3_notify = customtkinter.CTkEntry(newnotify_window, placeholder_text="Address Line 3")
    Addressline3_notify.grid(row=4,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    City_notify = customtkinter.CTkEntry(newnotify_window, placeholder_text="City")
    City_notify.grid(row=5,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    State_notify = customtkinter.CTkEntry(newnotify_window, placeholder_text="State")
    State_notify.grid(row=6,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    pincode_notify = customtkinter.CTkEntry(newnotify_window, placeholder_text="Pin code")
    pincode_notify.grid(row=7,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    Country_notify = customtkinter.CTkEntry(newnotify_window, placeholder_text="Country")
    Country_notify.grid(row=8,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    Phn_No_notify = customtkinter.CTkEntry(newnotify_window, placeholder_text="Phn No")
    Phn_No_notify.grid(row=9,column=1,padx=10,pady=5,sticky=W,columnspan=2)

    save_button = customtkinter.CTkButton(newnotify_window,text="Save",command=run_notify_script_and_save)
    save_button.grid(row=17,column=1,padx=10,pady=5,sticky=W)

    discard_button = customtkinter.CTkButton(newnotify_window,text="Discard",command=new_notify_window_destroy)
    discard_button.grid(row=19,column=1,padx=10,pady=5,sticky=W)

    newnotifystatus_label = customtkinter.CTkLabel(newnotify_window,text="",font=("Tahoma",12))
    newnotifystatus_label.grid(row=20,column=1,padx=10,pady=5,sticky=W)



#consignee details window_____________________________________________________________________________________________________________________________________________________________________________________________________

def consignee_search():
    consignee_search_window  = customtkinter.CTkToplevel(root)
    consignee_search_window.title("Comsignee Details")
    consignee_search_window.geometry("900x600")
    consignee_search_window.attributes(('-topmost'),True)
    consignee_search_window.grab_set()



#ADD NEW CONSIGNEE SECTION__________________________________________________________________________________________________________
    
    def newconsignee_window():
        newconsignee_window  = customtkinter.CTkToplevel(root)
        newconsignee_window.title("New Seller")
        newconsignee_window.geometry("400x800")
        newconsignee_window.attributes(('-topmost'),True)
        newconsignee_window.grab_set()



        def load_consignee_details():

            import pandas as pd
            

            # Load SELLER names from Excel file
            excel_file_path = 'I:/databases/Consignee details.xlsx'
            df = pd.read_excel(excel_file_path)
            company_names = df['Company name'].tolist()
            
            print(company_names)

            return df, company_names

        def update_consignee_names():
            global company_names
            _, company_names = load_consignee_details()
            listbox.delete(0, END)  # Clear existing items
            for name in company_names:
                listbox.insert(END, name)  # Insert new items
            print("variable updated")



        import pandas as pd

            # Destroy the window after a delay
        def newconsignee_window_destroy():    
            newconsignee_window.after(1500, newconsignee_window.destroy)

        def consigneedatarow():
            
            companyname = consignee_companyname_new.get()
            contactperson = consignee_contactperson_new.get()
            address1 = consignee_addressline1_new.get()
            address2 = consignee_Addressline2_new.get()
            address3 = consignee_Addressline3_new.get()
            City = consignee_city_new.get()
            State = consignee_state_new.get()
            pincode = consignee_pincode_new.get()
            Country = consignee_country_new.get()
            Phn_No = consignee_phn_no_new.get()


            # Return a dictionary with the values
            return {

                "Company name": companyname,
                "Contact person": contactperson,
                "Address line 1": address1,
                "Address line 2": address2,
                "Address line 3":address3,
                "City": City,
                "State": State,
                "Pin code": pincode,
                "Country": Country,
                "Phn No": Phn_No,

            }

        # Function to run the script and save data
        def run_script_and_save():
            try:
                import openpyxl
            except ImportError:
                print("Error", "openpyxl module not found. Make sure it's installed.")
                return

            def find_first_row_with_data(sheet, start_row):
                current_row = start_row
                while sheet.cell(row=current_row, column=1).value is None:
                    current_row -= 1
                return current_row

            # Main function to handle data mapping and writing to Excel using Pandas
            def main():
                data_dict = consigneedatarow()
                print(data_dict)

                # Define column mapping with correct capitalization
                column_mapping = {
                    "Company name": "Company name",
                    "Contact person": "Contact person",
                    "Address line 1": "Address line 1",
                    "Address line 2": "Address line 2",
                    "Address line 3": "Address line3",
                    "City": "City",
                    "State": "State",
                    "Pin code": "Pin code",
                    "Country": "Country",
                    "Phn No": "Phn No",

                }

                print("Column Mapping Keys:", column_mapping.keys())

                # Map data inputs to column names
                data_mapped = {column_mapping[key]: value for key, value in data_dict.items()if key in column_mapping}
                print("Mapped data:", data_mapped)

                # Create a DataFrame with the mapped data
                df = pd.DataFrame([data_mapped])

                # Load existing data from Excel
                try:
                    existing_data = pd.read_excel('I:/databases/Consignee details.xlsx')
                except FileNotFoundError:
                    print("ERROR FILE PATH IS INCORRECT")
                    existing_data = pd.DataFrame()

                # Concatenate existing data with new data
                updated_data = pd.concat([existing_data, df], ignore_index=True)

                # Write back to Excel
                updated_data.to_excel('I:/databases/Consignee details.xlsx', index=False,)
                print("Success", "Data entered successfully!")

                # Update label for user:
                newsellerstatus_label.configure(text=f'CONSIGNEE DETAILS ENTERED SUCCESSFULLY')

                # Reload the data for the rest of the code
                load_consignee_details()

                #update the combobox:
                update_consignee_names()            

                # Destroy the window after a delay
                newconsignee_window_destroy()

            main()


        consignee_companyname_new = customtkinter.CTkEntry(newconsignee_window, placeholder_text="Company name")
        consignee_companyname_new.grid(row=0,column=1,padx=10,pady=5,sticky=W,columnspan=2)
        
        consignee_contactperson_new = customtkinter.CTkEntry(newconsignee_window, placeholder_text="Contact Person details")
        consignee_contactperson_new.grid(row=1,column=1,padx=10,pady=5,sticky=W,columnspan=2)

        consignee_addressline1_new = customtkinter.CTkEntry(newconsignee_window, placeholder_text="Address Line 1")
        consignee_addressline1_new.grid(row=2,column=1,padx=10,pady=5,sticky=W,columnspan=2)

        consignee_Addressline2_new = customtkinter.CTkEntry(newconsignee_window, placeholder_text="Address Line 2")
        consignee_Addressline2_new.grid(row=3,column=1,padx=10,pady=5,sticky=W,columnspan=2)

        consignee_Addressline3_new = customtkinter.CTkEntry(newconsignee_window, placeholder_text="Address Line 3")
        consignee_Addressline3_new.grid(row=4,column=1,padx=10,pady=5,sticky=W,columnspan=2)

        consignee_city_new = customtkinter.CTkEntry(newconsignee_window, placeholder_text="City")
        consignee_city_new.grid(row=5,column=1,padx=10,pady=5,sticky=W,columnspan=2)

        consignee_state_new = customtkinter.CTkEntry(newconsignee_window, placeholder_text="State")
        consignee_state_new.grid(row=6,column=1,padx=10,pady=5,sticky=W,columnspan=2)

        consignee_pincode_new = customtkinter.CTkEntry(newconsignee_window, placeholder_text="Pin code")
        consignee_pincode_new.grid(row=7,column=1,padx=10,pady=5,sticky=W,columnspan=2)

        consignee_country_new = customtkinter.CTkEntry(newconsignee_window, placeholder_text="Country")
        consignee_country_new.grid(row=8,column=1,padx=10,pady=5,sticky=W,columnspan=2)

        consignee_phn_no_new = customtkinter.CTkEntry(newconsignee_window, placeholder_text="Phn No")
        consignee_phn_no_new.grid(row=9,column=1,padx=10,pady=5,sticky=W,columnspan=2)

        save_button = customtkinter.CTkButton(newconsignee_window,text="Save",command=run_script_and_save)
        save_button.grid(row=17,column=1,padx=10,pady=5,sticky=W)

        discard_button = customtkinter.CTkButton(newconsignee_window,text="Discard",command=newconsignee_window_destroy)
        discard_button.grid(row=18,column=1,padx=10,pady=5,sticky=W)

        newsellerstatus_label = customtkinter.CTkLabel(newconsignee_window,text="",font=("Tahoma",12))
        newsellerstatus_label.grid(row=19,column=1,padx=10,pady=5,sticky=W)






#CONSIGNEE SEARCH AND AUTO FILL
    # Update the listbox
    def update(data):
        # Clear the listbox
        listbox.delete(0, END)

        # Add company_name to listbox
        for item in data:
            listbox.insert(END, item)

    def newconsignee_window_destroy():    
        consignee_search_window .after(1500, consignee_search_window .destroy)

    def get_data():
        global consignee_data_checkbox
        consignee_data_checkbox = listbox.get()

    # Update entry box with listbox clicked
    def fillout(e):
        # Delete whatever is in the entry box
        consignee_search_entrybox.delete(0, END)

        # Add clicked list item to entry box
        consignee_search_entrybox.insert(0,'selected_option')

    # Create function to check entry vs listbox
    def check(e):
        # grab what was typed
        typed = consignee_search_entrybox.get()

        if typed == '':
            data = company_name
        else:
            data = []
            for item in company_name:
                if typed.lower() in item.lower():
                    data.append(item)

        # update our listbox with selected items
        update(data)				

    def diable_buyerconsignee_checkbox():
        consignee_check.configure(state="disabled")

    def list():
        # Load CONSIGNEE names from Excel file 
        excel_file_path = 'I:/databases/Consignee details.xlsx'
        df = pd.read_excel(excel_file_path)
        company_names = df['Company name'].tolist()
        print(company_names)
        return company_names


    def list_df():
        # Load CONSIGNEE names from Excel file 
        excel_file_path = 'I:/databases/Consignee details.xlsx'
        df = pd.read_excel(excel_file_path)
        df.fillna('NA', inplace=True)
        company_names = df['Company name'].tolist()
        print(company_names)
        return df
        


    def print_data_listbox1(selected_option):
        print(selected_option)
        update_labels(selected_option)



    def update_labels(selected_option):
        df = list_df()
        
        # Find corresponding entry in the DataFrame
        selected_row = df[df['Company name'] == selected_option].iloc[0]



        # Concatenate all rows, ignoring blank cells
        concatenated_details = ""
        for column, value in selected_row.items():
            if value != 'NA':
                concatenated_details = f"{selected_row['City']}, {selected_row['State']} {selected_row['Pin code']},{selected_row['Country']}"
        

        
        company_name = selected_row['Company name']
        contact_person = selected_row['Contact person']
        address1 = selected_row['Address line 1']
        address2 = selected_row['Address line 2']
        address3 = selected_row['Address line 3']
        citystatepincountry = concatenated_details
        phone = selected_row['Phn No']


        # Configure the label with concatenated details
        
        consignee_name_select.configure(text=f"Company Name:{company_name}", text_color="white")
        consignee_contactperson_select.configure(text=f"Contact person: {contact_person}", text_color="white")
        consignee_address1_select.configure(text=f"Address Line1: {address1}", text_color="white")
        consignee_address2_select.configure(text=f"Address Line2: {address2}", text_color="white")
        consignee_address3_select.configure(text=f"Address line3: {address3}", text_color="white")
        consignee_citystatepincountry_select.configure(text=citystatepincountry,text_color="white")
        consignee_phnno_select.configure(text=f"Phone No: {phone}")

        return company_name, contact_person,address1,address2,address3,citystatepincountry,phone



    def test():
        lol = listbox.get()
        print(lol)

        df = list_df()

        # Find corresponding entry in the DataFrame
        selected_row = df[df['Company name'] == lol].iloc[0]

        # Concatenate all rows, ignoring blank cells
        concatenated_details = ""
        for column, value in selected_row.items():
            if value != 'NA':
                concatenated_details = f"{selected_row['City']}, {selected_row['State']} {selected_row['Pin code']},{selected_row['Country']}"
        

        company_name = selected_row['Company name']
        contact_person = selected_row['Contact person']
        address1 = selected_row['Address line 1']
        address2 = selected_row['Address line 2']
        address3 = selected_row['Address line 3']
        citystatepincountry = concatenated_details
        country = selected_row['Country']
        phone = selected_row['Phn No']        

        print(company_name)
        print(contact_person )
        print(  address1 )
        print(address2 )
        print( address3 )
        print(  citystatepincountry )
        print(  country )
        print(        phone)


        # Open the Excel file for writing
        invoice_file_path = 'I:/invoice/Invoice.xlsx'
        invoice_wb = openpyxl.load_workbook(invoice_file_path)
        invoice_sheet = invoice_wb.active


        # Write data to the specified cells in the Invoice.xlsx sheet
        invoice_sheet['A22'] = (company_name )
        invoice_sheet['A23'] = (contact_person)
        invoice_sheet['A24'] = (address1)
        invoice_sheet['A25'] = ((address2 + address3))
        invoice_sheet['A26'] = (citystatepincountry)
        invoice_sheet['A27'] = (country)
        invoice_sheet['A28'] = (f"Phn No: {phone}")

        # Save the changes to the Excel file
        invoice_wb.save(invoice_file_path)

        # Convert all variables to strings
        company_name_str = str(company_name)
        contact_person_str = str(contact_person)
        address1_str = str(address1)
        address2_str = str(address2)
        address3_str = str(address3)
        citystatepincountry_str = str(citystatepincountry)
        phone_str = str(phone)

        # Concatenate all strings into one
        message = "Consignee: " + company_name_str + " Attn: " + contact_person_str + " ," + address1_str + " ," + address2_str + " ," + address3_str + " ," + citystatepincountry_str + " Phn No: " + phone_str

        # Configure consignee_label with the concatenated string
        consignee_label.configure(text=message)


        consignee_submit_status.configure(text=f'CONSIGNEE DETAILS ENTERED SUCCESSFULLY')

        print("data written successfully")

        diable_buyerconsignee_checkbox()
        
        
        newconsignee_window_destroy()

        get_data()

        


       


    #GUI
    # Create an entry box
    consignee_search_entrybox = customtkinter.CTkEntry(consignee_search_window,width=500,)
    consignee_search_entrybox.grid(row=5,column=1,padx=10,pady=5,sticky=W,columnspan=5)

    # Create a listbox
    listbox = CTkListbox(consignee_search_window, width=500,height=100,command = print_data_listbox1)
    listbox.grid( columnspan=5, padx=10, pady=0,row=7,column=1,sticky=W)

    # Create a list of pizza company_name
    company_name = list()

    # Add the company_name to our list
    update(company_name)


    # Create a binding on the entry box
    consignee_search_entrybox.bind("<KeyRelease>", check)


    my_button_newconsignee_details = customtkinter.CTkButton(consignee_search_window, text="Add New Consignee",command=newconsignee_window)
    my_button_newconsignee_details.grid(row=5, column=15, pady=5, columnspan=2, padx=10, sticky=W)


    consignee_name_select = customtkinter.CTkLabel(consignee_search_window,text="",font=("Tahoma",12))    
    consignee_name_select.grid( columnspan=5, padx=10, pady=5,row=20,column=1,sticky=W)

    consignee_contactperson_select = customtkinter.CTkLabel(consignee_search_window,text="",font=("Tahoma",12))    
    consignee_contactperson_select.grid( columnspan=5, padx=10, pady=5,row=21,column=1,sticky=W)

    consignee_address1_select = customtkinter.CTkLabel(consignee_search_window,text="",font=("Tahoma",12))    
    consignee_address1_select.grid( columnspan=5, padx=10, pady=5,row=22,column=1,sticky=W)

    consignee_address2_select = customtkinter.CTkLabel(consignee_search_window,text="",font=("Tahoma",12))    
    consignee_address2_select.grid( columnspan=5, padx=10, pady=5,row=23,column=1,sticky=W)

    consignee_address3_select = customtkinter.CTkLabel(consignee_search_window,text="",font=("Tahoma",12))    
    consignee_address3_select.grid( columnspan=5, padx=10, pady=5,row=24,column=1,sticky=W)

    consignee_citystatepincountry_select = customtkinter.CTkLabel(consignee_search_window,text="",font=("Tahoma",12))    
    consignee_citystatepincountry_select.grid( columnspan=5, padx=10, pady=5,row=25,column=1,sticky=W)

    consignee_phnno_select = customtkinter.CTkLabel(consignee_search_window,text="",font=("Tahoma",12))    
    consignee_phnno_select.grid( columnspan=5, padx=10, pady=5,row=26,column=1,sticky=W)


    newconsignee_details_submit = customtkinter.CTkButton(consignee_search_window, text="Submit",command=test)
    newconsignee_details_submit.grid(row=28, column=1, pady=5, columnspan=2, padx=10, sticky=W)

    
    consignee_submit_status =  customtkinter.CTkLabel(consignee_search_window,text="",font=("Tahoma",12))    
    consignee_submit_status.grid( columnspan=5, padx=10, pady=5,row=32,column=1,sticky=W)


    consignee_search_window.mainloop()


















#GUI ELEMENTS___________________________________________________________________________________________________________________________________________________________________


# Seller details
# Create combobox with increased width
df, company_names = load_seller_details()
my_combo = customtkinter.CTkComboBox(root, values=company_names, width=300)
my_combo.grid(row=1, column=0, pady=0, padx=10, columnspan=2, sticky=W)

# Create a Button
my_button = customtkinter.CTkButton(root, text="Submit Seller Details", command=choosecompany)
my_button.grid(row=1, column=5, pady=5, padx=10, sticky=W)

# Create a Button for New Seller details
my_button_new_buyer = customtkinter.CTkButton(root, text="Add New Seller", command=new_seller)
my_button_new_buyer.grid(row=1, column=10, pady=5, columnspan=3, padx=10,sticky=EW)

# Create output label
output_label = customtkinter.CTkLabel(root, text="", font=("Tahoma", 12))
output_label.grid(row=3, column=0, pady=5, columnspan=50, padx=10, sticky=W)


# Buyer details
# Create combobox with increased width for buyer details

df_buyer, buyer_names = load_buyer_details()
my_comboseller = customtkinter.CTkComboBox(root, values=buyer_names, width=int(300))
my_comboseller.grid(row=4, column=0, pady=0, padx=10, columnspan=2,sticky=W)

# Create a Button for buyer details
my_button_buyer = customtkinter.CTkButton(root, text="Submit Buyer Details", command=choosebuyer)
my_button_buyer.grid(row=4, column=5, pady=5, padx=10, sticky=W)

# Create a Button for buyer details
my_button_new_buyer = customtkinter.CTkButton(root, text="Add New Buyer", command=new_buyer)
my_button_new_buyer.grid(row=4, column=10, pady=5, columnspan=2,padx=10, sticky=EW )

# Create output label for buyer details
output_label_buyer = customtkinter.CTkLabel(root, text="", font=("Tahoma", 12))
output_label_buyer.grid(row=5, column=0, pady=5, columnspan=50, padx=10, sticky=W)


# Buyer same as Consignee
# Buer same as consignee checkbox
check_var = customtkinter.StringVar(value="off")

consignee_check = customtkinter.CTkCheckBox(root, text="Buyer the same as Consignee", variable=check_var, onvalue="on", offvalue="off", command=consignee_details)
consignee_check.grid(row=7, column=0, pady=5, columnspan=2,padx=10,sticky=W)

consignee_label = customtkinter.CTkLabel(root, text="")
consignee_label.grid(row=8, column=0, pady=5,columnspan=50,sticky=W,padx=10 )

# Create a Button for entering consignee details
my_button_consignee_details = customtkinter.CTkButton(root, text="Select Consignee", command=consignee_search)
my_button_consignee_details.grid(row=7, column=5, pady=5, columnspan=2)


# Notify party details
# Create combobox with increased width

df_notify, notify_company_names = load_Notify_details()
my_combo_notify = customtkinter.CTkComboBox(root, values=notify_company_names, width=300)
my_combo_notify.grid(row=9, column=0, pady=0, columnspan=2, padx=10,sticky=W)

# Create a Button
my_button_notify = customtkinter.CTkButton(root, text="Choose Notify Party", command=choosenotify)
my_button_notify.grid(row=9, column=5, pady=5, columnspan=2,sticky=W,padx=10)

# Create a Button for New Notify party
my_button_new_buyer = customtkinter.CTkButton(root, text="Add Notify Party", command=new_notify)
my_button_new_buyer.grid(row=9, column=10, pady=5, columnspan=2,sticky=EW,padx=10)

# Create output label
output_label_notify = customtkinter.CTkLabel(root, text="", font=("Tahoma", 12))
output_label_notify.grid(row=10, column=0, pady=5, padx=10,columnspan=50,sticky=W)


#Consignee as second notify party
check_var2 = customtkinter.StringVar(value="off")

secondconsignee_check = customtkinter.CTkCheckBox(root, text="Consignee as second notify party", variable=check_var2, onvalue="on", offvalue="off", command=secondconsignee_details)
secondconsignee_check.grid(row=11, column=0, pady=5, padx=10,sticky=W)

secondconsignee_label = customtkinter.CTkLabel(root, text="")
secondconsignee_label.grid(row=12, column=0, pady=5,columnspan=50,padx=10,sticky=W)

#Port of Delivery
portofdelivery = customtkinter.CTkEntry(root, placeholder_text="Enter Port of Delivery", font=("Tahoma", 12))
portofdelivery.grid(row=17, column=0, pady=5,padx=10,sticky=W)

portofdeliverybutton = customtkinter.CTkButton(root, text="Submit", command=submit_and_country)
portofdeliverybutton.grid(row=17, column=5, pady=5, columnspan=2,padx=10,sticky=W)

portofdeliverylabel = customtkinter.CTkLabel(root, text="")
portofdeliverylabel.grid(row=19, column=0, pady=5,padx=10,sticky=W)

#Incoterms
incoterms_var = customtkinter.StringVar(value="other")

incoterms_1 = customtkinter.CTkRadioButton(root,text = "CIF",value="Yes",variable=incoterms_var)
incoterms_1.grid(row=20,column=0,padx=10,sticky=W)


incoterms_2 = customtkinter.CTkRadioButton(root,text = "FOB",value="No",variable=incoterms_var)
incoterms_2.grid(row=20,column=5,padx=10,sticky=W)

incoterms_button = customtkinter.CTkButton(root, text="Submit", command=get_rad)
incoterms_button.grid(row=20,column=15,padx=10,sticky=W)

incoterms_label = customtkinter.CTkLabel(root,text="",font=("Tahoma", 12))
incoterms_label.grid(row=21,column=0,padx=10,sticky=W)



#Shipping mode
shippingmode_var = customtkinter.IntVar(value="1")

shippingmode_seafreight = customtkinter.CTkRadioButton(root,text = "SEA FREIGHT",value="2",variable=shippingmode_var)
shippingmode_seafreight.grid(row=27,column=0,padx=10,sticky=W)

shippingmode_airfreight = customtkinter.CTkRadioButton(root,text = "AIR FREIGHT",value="3",variable=shippingmode_var)
shippingmode_airfreight.grid(row=27,column=5,padx=10,sticky=W)

shippingmode_courier = customtkinter.CTkRadioButton(root, text = "COURIER",value= "4",variable=shippingmode_var)
shippingmode_courier.grid(row=27,column=10,padx=10,sticky=W)

shippingmode_button = customtkinter.CTkButton(root,text="Submit",command=shippingtermpayment)
shippingmode_button.grid(row=27,column=15,padx=10,sticky=W)

shippingmode_label = customtkinter.CTkLabel(root,text="",font=("Tahoma",12))
shippingmode_label.grid(row=30,column=0,padx=10,sticky=W)



#SHIPPING MARKS
shippingmarks = customtkinter.CTkEntry(root,placeholder_text="Enter Shipping Marks", font=("Tahoma", 12))
shippingmarks.grid(row=31,column=0,padx=10,pady=5,sticky=W)

shippingmarks_button = customtkinter.CTkButton(root, text="Submit",command=shippingmarksinput)
shippingmarks_button.grid(row=31, column=5, pady=5, columnspan=2,padx=10,sticky=W)

marks_label = customtkinter.CTkLabel(root, text="")
marks_label.grid(row=32, column=0, pady=5,padx=10,sticky=W)


#TC Status
getdatatest = customtkinter.CTkButton(root, text="check get data function",command=getdatatest1)
getdatatest.grid(row=40, column=5, pady=5, columnspan=2,padx=10,sticky=W)


root.mainloop()
